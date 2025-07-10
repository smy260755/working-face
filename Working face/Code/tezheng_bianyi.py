import pandas as pd
import math
from openpyxl import Workbook
import numpy as np
import time

# 计算点到直线的距离
def point_to_line_distance(point_a, point_b, point_c):
    if point_b[0] == point_c[0]:
        return 9999999
    slope = (point_b[1] - point_c[1]) / (point_b[0] - point_c[0])
    intercept = point_b[1] - slope * point_b[0]
    distance = abs(slope * point_a[0] - point_a[1] + intercept) / math.sqrt(1 + pow(slope, 2))
    return distance

# 计算变异系数
def calculate_cv(data):
    mean = np.mean(data)
    std = np.std(data)
    return std / mean if mean else 0

# 动态计算阈值
def calculate_dynamic_threshold(speeds, base_threshold):
    speed_cv = calculate_cv(speeds)
    return base_threshold * (1 + speed_cv)

# 执行道格拉斯-普克算法
def douglas_peucker_with_dynamic_threshold(points, base_threshold, lowerLimit=4, ceiling=None):
    def diluting(point_list, threshold, qualify_list, disqualify_list):
        if len(point_list) < 3:
            qualify_list.extend(point_list[::-1])
        else:
            max_distance_index, max_distance = 0, 0
            for index, point in enumerate(point_list):
                if index in [0, len(point_list) - 1]:
                    continue
                distance = point_to_line_distance(point, point_list[0], point_list[-1])
                if distance > max_distance:
                    max_distance_index = index
                    max_distance = distance

            if max_distance < threshold:
                qualify_list.append(point_list[-1])
                qualify_list.append(point_list[0])
            else:
                sequence_a = point_list[:max_distance_index]
                sequence_b = point_list[max_distance_index:]

                for sequence in [sequence_a, sequence_b]:
                    if len(sequence) < 3 and sequence == sequence_b:
                        qualify_list.extend(sequence[::-1])
                    else:
                        disqualify_list.append(sequence)
        return qualify_list, disqualify_list

    def get_qualify_list(point_list, threshold):
        qualify_list = []
        disqualify_list = []

        qualify_list, disqualify_list = diluting(point_list, threshold, qualify_list, disqualify_list)
        while len(disqualify_list) > 0:
            qualify_list, disqualify_list = diluting(disqualify_list.pop(), threshold, qualify_list, disqualify_list)

        return qualify_list

    if len(points) < 5:
        return points

    start_time = time.time()

    # 计算速度的变异系数
    speeds = [point[2] for point in points]
    threshold = calculate_dynamic_threshold(speeds, base_threshold)

    result = get_qualify_list(points, threshold)

    end_time = time.time()

    original_length = len(points)
    simplified_length = len(result)
    distortion = (original_length - simplified_length) / original_length

    average_error = 0
    max_error = 0
    for i in range(len(points)):
        original_point = points[i]
        simplified_point = result[i] if i < len(result) else None
        if simplified_point:
            error = abs(original_point[0] - simplified_point[0]) + abs(original_point[1] - simplified_point[1])
            average_error += error
            max_error = max(max_error, error)
    average_error /= len(points)

    compressed_point_count = len(result)
    compression_time = end_time - start_time

    if len(result) < lowerLimit:
        while len(result) < lowerLimit:
            threshold *= 0.9
            result = get_qualify_list(points, threshold)

    if ceiling and len(result) > ceiling:
        while len(result) > ceiling:
            threshold *= 1.1
            result = get_qualify_list(points, threshold)

    if len(result) > len(points):
        return points

    return result, distortion, average_error, max_error, compressed_point_count, compression_time

# 读取 Excel 文件数据
data = pd.read_excel('../data/604_s03.xlsx')

# 提取 x 坐标 Out1, y 坐标 Out2 和速度 v
Out1 = data['Out1']
Out2 = data['Out2']
v = data['v']
print(len(Out1))
# 将坐标数据转换为点的列表，按照 mdt 降序排列
points = [(Out1[i], Out2[i], v[i]) for i in sorted(range(len(Out1)), key=lambda x: data['mdt'][x], reverse=True)]

# 执行道格拉斯-普克算法
base_threshold = 0.06  # 基础阈值
simplified_points, distortion, average_error, max_error, compressed_point_count, compression_time = douglas_peucker_with_dynamic_threshold(points, base_threshold)

# 创建工作簿
wb = Workbook()
ws = wb.active

# 设置列标题
ws['A1'] = 'mdt'
ws['B1'] = 'Out1'
ws['C1'] = 'Out2'
ws['D1'] = 'v'

# 写入抽稀后的数据
for i, point in enumerate(simplified_points):
    ws.cell(row=i + 2, column=1, value=data['mdt'][list(points).index(point)])
    ws.cell(row=i + 2, column=2, value=point[0])
    ws.cell(row=i + 2, column=3, value=point[1])
    ws.cell(row=i + 2, column=4, value=point[2])

# 保存工作簿
wb.save('../data/604_03th.xlsx')

# 读取保存的工作簿
data = pd.read_excel('../data/604_03th.xlsx')
time = data['mdt']
print(len(time))
print("平均误差:", average_error)
print("最大误差:", max_error)
print("压缩后的特征点数量:", compressed_point_count)
print("压缩时间:", compression_time)


import matplotlib.pyplot as plt

# 绘制原始点图
plt.figure(figsize=(10, 6))
plt.scatter(data['Out1'], data['Out2'], c='blue', label='origin point', s=20)
plt.title('origin points')
plt.xlabel('X')
plt.ylabel('Y')
plt.legend()
plt.savefig('../data/original_points.png', dpi=300)
plt.show()

# 绘制压缩点图
plt.figure(figsize=(10, 6))
simplified_data = pd.read_excel('../data/604_03th.xlsx')
plt.scatter(simplified_data['Out1'], simplified_data['Out2'], c='red', label='Compression point', s=20)
plt.title('Compression points')
plt.xlabel('X')
plt.ylabel('Y')
plt.legend()
plt.savefig('../data/compressed_points.png', dpi=300)
plt.show()